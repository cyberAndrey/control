# -*- coding: utf-8 -*-
import calendar
import datetime
import io
import openpyxl
import os.path
import xlsxwriter

from flask import render_template, redirect, request

from app import app, db
from app.models import Key, Staff, Guest


class NameFile:
    t = datetime.datetime.now()
    watchmanFile = str(t.year) + "_" + str(t.month) + "_watchman" + ".xlsx"
    teacherFile = str(t.year) + "_" + str(t.month) + "_teachers" + ".xlsx"
    isOpen = False


@app.route('/')
@app.route('/index')
def index():
    guests = Guest.query.all()
    staff = Staff.query.all()
    keys = Key.query.all()
    return render_template('index.html', title='Кванториум', guests=guests, staff=staff, keys=keys)


@app.route('/addNote', methods=['POST'])
def addNote():
    if 'come' in request.form:
        t = datetime.datetime.now()
        filename = str(t.year) + "_" + str(t.month) + "_teachers" + ".xlsx"
        NameFile.teacherFile = filename
        createTableIfNotExist(filename)
        name = request.form.get('personName')
        temperature = request.form.get('temperature')
        notes = [str(t.month) + "_" + str(t.day), name, '+', temperature]
        addNoteInTable(NameFile.teacherFile, notes, 'A')
    else:
        print('start')
        pos = findRow(request.form.get('personName'))
        print(pos)
        if pos is not None:
            wb = openpyxl.load_workbook(NameFile.teacherFile)
            sheet = wb['sheet']
            temperature = request.form.get('temperature')
            notes = ['+', temperature]
            col = ord('E')
            for note in notes:
                sheet[chr(col) + str(pos)] = note
                col += 1
            wb.save(NameFile.teacherFile)
    return redirect('/index')


@app.route('/openShift', methods=['POST'])
def openShift():
    if 'open' in request.form and not NameFile.isOpen:
        t = datetime.datetime.now()
        filename = str(t.year) + "_" + str(t.month) + "_watchman" + ".xlsx"
        NameFile.watchmanFile = filename
        createTableIfNotExist(filename)
        name = request.form.get('personName')
        temperature = request.form.get('temperature')
        time = str(t.hour) + ":" + str(t.minute)
        notes = [str(t.year) + "_" + str(t.month), time, name, temperature]
        addNoteInTable(NameFile.watchmanFile, notes, 'A')
        NameFile.isOpen = True
    elif 'close' in request.form:
        name = request.form.get('personName')
        temperature = request.form.get('temperature')
        notes = [name, temperature]
        addNoteInTable(NameFile.watchmanFile, notes, 'E')
        NameFile.isOpen = False
    return redirect('/index')


@app.route('/addGuest', methods=['POST'])
def addGuest():
    name = request.form.get('gCome')
    g = Guest(name=name)
    db.session.add(g)
    db.session.commit()
    f = open('guests.txt', 'a', encoding="utf-8")
    f.write('Гость ' + name + ' пришёл в ' + str(datetime.datetime.now().date()) +
            ' ' + str(datetime.datetime.now().time()) + '\n')
    f.close()
    return redirect('/index')


@app.route('/deleteGuest', methods=['POST'])
def deleteGuest():
    name = request.form.get('guests')
    g = Guest.query.filter(Guest.name == name).first()
    db.session.delete(g)
    db.session.commit()
    f = open('guests.txt', 'a', encoding="utf-8")
    f.write('Гость ' + name + ' ушёл в ' + str(datetime.datetime.now().date()) +
            ' ' + str(datetime.datetime.now().time()) + '\n')
    f.close()
    return redirect('/index')


@app.route('/takeKey', methods=['POST'])
def takeKey():
    name = request.form.get('personName')
    kvantum = request.form.get('kvantum')
    k = Key.query.filter(Key.kvantum == kvantum).first()
    k.taken = True
    k.who = name
    db.session.commit()
    return redirect('/index')


@app.route('/backKey', methods=['POST'])
def backKey():
    kvantum = request.form.get('back')
    k = Key.query.filter(Key.kvantum == kvantum).first()
    k.taken = False
    k.who = None
    db.session.commit()
    return redirect('/index')


def createTableIfNotExist(name):
    headers = ['дата', 'ФИО', 'прибыл', 't по прибытию', 'убыл', 't по убытию'] if 'teachers' in name \
        else ['дата передачи смены', 'время передачи смены', 'Сдал ФИО', 't-тела', 'Принял ФИО', 't-тела']
    if not os.path.exists(name):
        workbook = xlsxwriter.Workbook(name)
        worksheet = workbook.add_worksheet('sheet')
        fillTable(worksheet, headers)
        workbook.close()
        return workbook.filename
    return name


def addNoteInTable(file, notes, col):
    wb = openpyxl.load_workbook(file)
    sheet = wb['sheet']
    if col == 'A':
        str_in_table = sheet.max_row + 1
    else:
        str_in_table = sheet.max_row
    col = ord(col)
    for note in notes:
        sheet[chr(col) + str(str_in_table)] = note
        col += 1
    wb.save(file)


def findRow(name):
    print(NameFile.teacherFile)
    wb = openpyxl.load_workbook(NameFile.teacherFile)
    sheet = wb['sheet']
    i = sheet.max_row
    while True:
        row = 'B' + str(i)
        if sheet[row].value == name:
            return i
        i -= 1
        if i == 0:
            return None


def findDate(sheet):
    t = datetime.date.today()
    need_date = str(t.day) + "." + str(t.month)
    col = ord('A')
    while True:
        if sheet[chr(col) + str(1)].value == need_date:
            return chr(col)
        col += 1
        if col == 'Z':
            return None  # Исправить костыль


def fillTable(worksheet, headers):
    column = 0
    for i in headers:
        worksheet.write(0, column, i)
        column += 1


@app.route('/aboutKeys', methods=['GET'])
def aboutKeys():
    keys = Key.query.all()
    return render_template('keys.html', keys=keys)
